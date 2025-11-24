"""
Export utilities for Excel and PDF generation
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from datetime import datetime, date, timedelta
from typing import List, Dict
import io


def generate_date_range(start_date: date, end_date: date) -> List[date]:
    """Generate a list of dates between start and end"""
    dates = []
    current = start_date
    while current <= end_date:
        dates.append(current)
        current += timedelta(days=1)
    return dates


def is_weekend(check_date: date) -> bool:
    """Check if a date is weekend"""
    return check_date.weekday() >= 5


def export_to_excel(employees_data: List[Dict], start_date: date, end_date: date) -> bytes:
    """
    Export calendar to Excel
    
    employees_data format:
    [
        {
            'name': 'Employee Name',
            'role': 'Role',
            'team': 'Team Name',
            'vacations': [{'start_date': date, 'end_date': date, 'type': 'vacation'}]
        }
    ]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Calendario de Vacaciones"
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    weekend_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    vacation_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    sick_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    personal_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Headers
    ws['A1'] = "Nombre"
    ws['B1'] = "Rol"
    ws['C1'] = "Equipo"
    
    # Aplicar estilo a headers fijos
    for col in ['A', 'B', 'C']:
        ws[f'{col}1'].fill = header_fill
        ws[f'{col}1'].font = header_font
        ws[f'{col}1'].alignment = center_align
        ws[f'{col}1'].border = border
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    
    # Generar fechas
    dates = generate_date_range(start_date, end_date)
    
    # Fechas como headers
    for col_idx, date_obj in enumerate(dates, start=4):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = date_obj.strftime("%d-%b")
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = 10
    
    # Columna de totales
    total_col = len(dates) + 4
    ws.cell(row=1, column=total_col, value="Total Días").fill = header_fill
    ws.cell(row=1, column=total_col).font = header_font
    ws.cell(row=1, column=total_col).alignment = center_align
    ws.cell(row=1, column=total_col).border = border
    ws.column_dimensions[ws.cell(row=1, column=total_col).column_letter].width = 12
    
    # Datos de empleados
    for row_idx, emp_data in enumerate(employees_data, start=2):
        ws.cell(row=row_idx, column=1, value=emp_data['name']).border = border
        ws.cell(row=row_idx, column=2, value=emp_data.get('role', '')).border = border
        ws.cell(row=row_idx, column=3, value=emp_data.get('team', '')).border = border
        
        # Crear set de fechas de vacaciones
        vacation_dates = set()
        vacation_types = {}
        
        for vacation in emp_data.get('vacations', []):
            current = vacation['start_date']
            while current <= vacation['end_date']:
                vacation_dates.add(current)
                vacation_types[current] = vacation.get('type', 'vacation')
                current += timedelta(days=1)
        
        total_days = 0
        
        for col_idx, date_obj in enumerate(dates, start=4):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            if date_obj in vacation_dates:
                cell.value = "X"
                vac_type = vacation_types.get(date_obj, 'vacation')
                
                if vac_type == 'sick':
                    cell.fill = sick_fill
                elif vac_type == 'personal':
                    cell.fill = personal_fill
                else:
                    cell.fill = vacation_fill
                
                cell.alignment = center_align
                total_days += 1
            elif is_weekend(date_obj):
                cell.fill = weekend_fill
            
            cell.border = border
        
        # Total
        cell = ws.cell(row=row_idx, column=total_col)
        cell.value = total_days
        cell.alignment = center_align
        cell.border = border
    
    # Guardar en bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def export_to_pdf(employees_data: List[Dict], start_date: date, end_date: date, 
                  title: str = "Calendario de Vacaciones") -> bytes:
    """
    Export calendar to PDF
    
    employees_data format: same as export_to_excel
    """
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#4472C4'),
        spaceAfter=30,
        alignment=1  # Center
    )
    title_para = Paragraph(title, title_style)
    elements.append(title_para)
    
    subtitle = Paragraph(
        f"Período: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}",
        styles['Normal']
    )
    elements.append(subtitle)
    elements.append(Spacer(1, 0.5*cm))
    
    # Generar fechas
    dates = generate_date_range(start_date, end_date)
    
    # Preparar datos de la tabla
    table_data = []
    
    # Headers
    headers = ["Nombre", "Rol", "Equipo"] + [d.strftime("%d-%b") for d in dates] + ["Total"]
    table_data.append(headers)
    
    # Datos de empleados
    for emp_data in employees_data:
        row = [emp_data['name'], emp_data.get('role', ''), emp_data.get('team', '')]
        
        # Crear set de fechas de vacaciones
        vacation_dates = set()
        for vacation in emp_data.get('vacations', []):
            current = vacation['start_date']
            while current <= vacation['end_date']:
                vacation_dates.add(current)
                current += timedelta(days=1)
        
        total_days = 0
        for date_obj in dates:
            if date_obj in vacation_dates:
                row.append("X")
                total_days += 1
            else:
                row.append("")
        
        row.append(str(total_days))
        table_data.append(row)
    
    # Crear tabla
    table = Table(table_data)
    
    # Estilo de la tabla
    table_style = TableStyle([
        # Headers
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Datos
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
        
        # Columna de totales
        ('BACKGROUND', (-1, 0), (-1, -1), colors.HexColor('#E7E6E6')),
        ('FONTNAME', (-1, 1), (-1, -1), 'Helvetica-Bold'),
    ])
    
    # Resaltar fines de semana
    for col_idx, date_obj in enumerate(dates, start=3):
        if is_weekend(date_obj):
            table_style.add('BACKGROUND', (col_idx, 1), (col_idx, -1), colors.HexColor('#E7E6E6'))
    
    # Resaltar vacaciones
    for row_idx, emp_data in enumerate(employees_data, start=1):
        vacation_dates = set()
        for vacation in emp_data.get('vacations', []):
            current = vacation['start_date']
            while current <= vacation['end_date']:
                vacation_dates.add(current)
                current += timedelta(days=1)
        
        for col_idx, date_obj in enumerate(dates, start=3):
            if date_obj in vacation_dates:
                table_style.add('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), 
                              colors.HexColor('#92D050'))
    
    table.setStyle(table_style)
    elements.append(table)
    
    # Leyenda
    elements.append(Spacer(1, 0.5*cm))
    legend_style = ParagraphStyle(
        'Legend',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.grey
    )
    legend = Paragraph(
        "<b>Leyenda:</b> X = Vacaciones | Gris = Fin de semana | Verde = Día de vacaciones",
        legend_style
    )
    elements.append(legend)
    
    doc.build(elements)
    output.seek(0)
    return output.getvalue()
