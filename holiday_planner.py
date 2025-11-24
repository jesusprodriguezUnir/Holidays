"""
Holiday Planner Application
Gestión de vacaciones del equipo durante el período navideño 2025
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta
import json
import os
from typing import List, Dict, Set
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm


class TeamMember:
    """Representa un miembro del equipo con sus días de vacaciones"""
    
    def __init__(self, name: str, role: str = "", vacation_days: Set[str] = None):
        self.name = name
        self.role = role
        self.vacation_days = vacation_days if vacation_days else set()
    
    def add_vacation_day(self, date_str: str):
        """Añade un día de vacaciones"""
        self.vacation_days.add(date_str)
    
    def remove_vacation_day(self, date_str: str):
        """Elimina un día de vacaciones"""
        self.vacation_days.discard(date_str)
    
    def is_on_vacation(self, date_str: str) -> bool:
        """Verifica si está de vacaciones en una fecha"""
        return date_str in self.vacation_days
    
    def to_dict(self) -> Dict:
        """Convierte a diccionario para serialización"""
        return {
            "name": self.name,
            "role": self.role,
            "vacation_days": list(self.vacation_days)
        }
    
    @staticmethod
    def from_dict(data: Dict) -> 'TeamMember':
        """Crea un TeamMember desde un diccionario"""
        return TeamMember(
            name=data["name"],
            role=data.get("role", ""),
            vacation_days=set(data.get("vacation_days", []))
        )


class HolidayPlanner:
    """Gestor principal del planificador de vacaciones"""
    
    def __init__(self, start_date: datetime, end_date: datetime):
        self.start_date = start_date
        self.end_date = end_date
        self.team_members: List[TeamMember] = []
        self.data_file = "holiday_data.json"
        
        # Generar lista de fechas
        self.dates = []
        current = start_date
        while current <= end_date:
            self.dates.append(current)
            current += timedelta(days=1)
    
    def add_member(self, member: TeamMember):
        """Añade un miembro al equipo"""
        self.team_members.append(member)
    
    def remove_member(self, name: str):
        """Elimina un miembro del equipo"""
        self.team_members = [m for m in self.team_members if m.name != name]
    
    def get_member(self, name: str) -> TeamMember:
        """Obtiene un miembro por nombre"""
        for member in self.team_members:
            if member.name == name:
                return member
        return None
    
    def is_weekend(self, date: datetime) -> bool:
        """Verifica si una fecha es fin de semana"""
        return date.weekday() >= 5  # 5=Sábado, 6=Domingo
    
    def save_data(self):
        """Guarda los datos en archivo JSON"""
        data = {
            "team_members": [m.to_dict() for m in self.team_members]
        }
        with open(self.data_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    
    def load_data(self):
        """Carga los datos desde archivo JSON"""
        if os.path.exists(self.data_file):
            try:
                with open(self.data_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.team_members = [TeamMember.from_dict(m) for m in data.get("team_members", [])]
                return True
            except Exception as e:
                print(f"Error cargando datos: {e}")
                return False
        return False
    
    def export_to_excel(self, filename: str):
        """Exporta el calendario a Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Vacaciones Navidad 2025"
        
        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        weekend_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        vacation_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Headers
        ws['A1'] = "Nombre"
        ws['B1'] = "Rol"
        
        # Fechas como headers
        for col_idx, date in enumerate(self.dates, start=3):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = date.strftime("%d-%b")
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
            ws.column_dimensions[cell.column_letter].width = 10
        
        # Aplicar estilo a headers de nombre y rol
        ws['A1'].fill = header_fill
        ws['A1'].font = header_font
        ws['A1'].alignment = center_align
        ws['A1'].border = border
        ws['B1'].fill = header_fill
        ws['B1'].font = header_font
        ws['B1'].alignment = center_align
        ws['B1'].border = border
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        
        # Datos de miembros
        for row_idx, member in enumerate(self.team_members, start=2):
            ws.cell(row=row_idx, column=1, value=member.name).border = border
            ws.cell(row=row_idx, column=2, value=member.role).border = border
            
            for col_idx, date in enumerate(self.dates, start=3):
                cell = ws.cell(row=row_idx, column=col_idx)
                date_str = date.strftime("%Y-%m-%d")
                
                if member.is_on_vacation(date_str):
                    cell.value = "X"
                    cell.fill = vacation_fill
                    cell.alignment = center_align
                elif self.is_weekend(date):
                    cell.fill = weekend_fill
                
                cell.border = border
        
        # Columna de totales
        total_col = len(self.dates) + 3
        ws.cell(row=1, column=total_col, value="Total Días").fill = header_fill
        ws.cell(row=1, column=total_col).font = header_font
        ws.cell(row=1, column=total_col).alignment = center_align
        ws.cell(row=1, column=total_col).border = border
        ws.column_dimensions[ws.cell(row=1, column=total_col).column_letter].width = 12
        
        for row_idx, member in enumerate(self.team_members, start=2):
            cell = ws.cell(row=row_idx, column=total_col)
            cell.value = len(member.vacation_days)
            cell.alignment = center_align
            cell.border = border
        
        wb.save(filename)
    
    def export_to_pdf(self, filename: str):
        """Exporta el calendario a PDF"""
        doc = SimpleDocTemplate(filename, pagesize=landscape(A4))
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#4472C4'),
            spaceAfter=30,
            alignment=1  # Center
        )
        title = Paragraph("Cuadrante de Vacaciones - Navidad 2025", title_style)
        elements.append(title)
        
        subtitle = Paragraph(
            f"Período: {self.start_date.strftime('%d/%m/%Y')} - {self.end_date.strftime('%d/%m/%Y')}",
            styles['Normal']
        )
        elements.append(subtitle)
        elements.append(Spacer(1, 0.5*cm))
        
        # Preparar datos de la tabla
        table_data = []
        
        # Headers
        headers = ["Nombre", "Rol"] + [date.strftime("%d-%b") for date in self.dates] + ["Total"]
        table_data.append(headers)
        
        # Datos de miembros
        for member in self.team_members:
            row = [member.name, member.role]
            for date in self.dates:
                date_str = date.strftime("%Y-%m-%d")
                if member.is_on_vacation(date_str):
                    row.append("X")
                else:
                    row.append("")
            row.append(str(len(member.vacation_days)))
            table_data.append(row)
        
        # Crear tabla
        table = Table(table_data)
        
        # Estilo de la tabla
        table_style = TableStyle([
            # Headers
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Datos
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
            
            # Columna de totales
            ('BACKGROUND', (-1, 0), (-1, -1), colors.HexColor('#E7E6E6')),
            ('FONTNAME', (-1, 1), (-1, -1), 'Helvetica-Bold'),
        ])
        
        # Resaltar fines de semana
        for col_idx, date in enumerate(self.dates, start=2):
            if self.is_weekend(date):
                table_style.add('BACKGROUND', (col_idx, 1), (col_idx, -1), colors.HexColor('#E7E6E6'))
        
        # Resaltar vacaciones
        for row_idx, member in enumerate(self.team_members, start=1):
            for col_idx, date in enumerate(self.dates, start=2):
                date_str = date.strftime("%Y-%m-%d")
                if member.is_on_vacation(date_str):
                    table_style.add('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#92D050'))
        
        table.setStyle(table_style)
        elements.append(table)
        
        # Leyenda
        elements.append(Spacer(1, 0.5*cm))
        legend_style = ParagraphStyle(
            'Legend',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.grey
        )
        legend = Paragraph(
            "<b>Leyenda:</b> X = Vacaciones | Gris = Fin de semana | Verde = Día de vacaciones",
            legend_style
        )
        elements.append(legend)
        
        doc.build(elements)


class HolidayPlannerGUI:
    """Interfaz gráfica de la aplicación"""
    
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Planificador de Vacaciones - Navidad 2025")
        self.root.geometry("1200x700")
        
        # Crear planificador
        start_date = datetime(2025, 12, 22)
        end_date = datetime(2026, 1, 7)
        self.planner = HolidayPlanner(start_date, end_date)
        
        # Cargar datos o inicializar con datos por defecto
        if not self.planner.load_data():
            self.initialize_default_data()
        
        self.selected_member = None
        
        # Crear interfaz
        self.create_widgets()
        self.refresh_grid()
    
    def initialize_default_data(self):
        """Inicializa con los datos proporcionados por el usuario"""
        # Jesús: 22 dic - 7 ene
        jesus_days = set()
        current = datetime(2025, 12, 22)
        end = datetime(2026, 1, 7)
        while current <= end:
            jesus_days.add(current.strftime("%Y-%m-%d"))
            current += timedelta(days=1)
        
        jesus = TeamMember("Jesús", "Team Lead", jesus_days)
        self.planner.add_member(jesus)
        
        # Adrian: 22 dic - 7 ene (igual que Jesús)
        adrian = TeamMember("Adrian", "Developer", jesus_days.copy())
        self.planner.add_member(adrian)
        
        # Felix: 24, 29, 30, 31 dic
        felix_days = {
            "2025-12-24",
            "2025-12-29",
            "2025-12-30",
            "2025-12-31"
        }
        felix = TeamMember("Felix", "Developer", felix_days)
        self.planner.add_member(felix)
        
        # Otros miembros sin vacaciones definidas
        self.planner.add_member(TeamMember("Alan", "Developer"))
        self.planner.add_member(TeamMember("Erik", "Developer"))
        self.planner.add_member(TeamMember("Cesar", "Developer"))
        
        self.planner.save_data()
    
    def create_widgets(self):
        """Crea los widgets de la interfaz"""
        # Frame superior con título y botones
        top_frame = ttk.Frame(self.root, padding="10")
        top_frame.pack(fill=tk.X)
        
        title_label = ttk.Label(
            top_frame,
            text="Cuadrante de Vacaciones - Navidad 2025",
            font=('Arial', 16, 'bold')
        )
        title_label.pack(side=tk.LEFT)
        
        # Botones de acción
        btn_frame = ttk.Frame(top_frame)
        btn_frame.pack(side=tk.RIGHT)
        
        ttk.Button(btn_frame, text="💾 Guardar", command=self.save_data).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="📊 Exportar Excel", command=self.export_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="📄 Exportar PDF", command=self.export_pdf).pack(side=tk.LEFT, padx=5)
        
        # Frame principal con grid y panel lateral
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Frame para el grid (izquierda)
        grid_frame = ttk.Frame(main_frame)
        grid_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Canvas y scrollbars para el grid
        canvas = tk.Canvas(grid_frame)
        v_scrollbar = ttk.Scrollbar(grid_frame, orient=tk.VERTICAL, command=canvas.yview)
        h_scrollbar = ttk.Scrollbar(grid_frame, orient=tk.HORIZONTAL, command=canvas.xview)
        
        self.grid_container = ttk.Frame(canvas)
        
        canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        canvas.create_window((0, 0), window=self.grid_container, anchor=tk.NW)
        self.grid_container.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        
        # Panel lateral (derecha)
        side_panel = ttk.LabelFrame(main_frame, text="Gestión de Equipo", padding="10")
        side_panel.pack(side=tk.RIGHT, fill=tk.Y, padx=(10, 0))
        
        # Sección de miembro seleccionado
        ttk.Label(side_panel, text="Miembro seleccionado:", font=('Arial', 10, 'bold')).pack(anchor=tk.W)
        self.selected_label = ttk.Label(side_panel, text="Ninguno", foreground="gray")
        self.selected_label.pack(anchor=tk.W, pady=(0, 10))
        
        # Añadir nuevo miembro
        ttk.Label(side_panel, text="Nuevo miembro:", font=('Arial', 10, 'bold')).pack(anchor=tk.W, pady=(10, 5))
        
        ttk.Label(side_panel, text="Nombre:").pack(anchor=tk.W)
        self.new_name_entry = ttk.Entry(side_panel, width=25)
        self.new_name_entry.pack(fill=tk.X, pady=(0, 5))
        
        ttk.Label(side_panel, text="Rol:").pack(anchor=tk.W)
        self.new_role_entry = ttk.Entry(side_panel, width=25)
        self.new_role_entry.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(side_panel, text="➕ Añadir Miembro", command=self.add_member).pack(fill=tk.X)
        
        # Editar miembro existente
        ttk.Separator(side_panel, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=15)
        
        ttk.Label(side_panel, text="Editar miembro:", font=('Arial', 10, 'bold')).pack(anchor=tk.W, pady=(0, 5))
        
        ttk.Label(side_panel, text="Nombre:").pack(anchor=tk.W)
        self.edit_name_entry = ttk.Entry(side_panel, width=25)
        self.edit_name_entry.pack(fill=tk.X, pady=(0, 5))
        
        ttk.Label(side_panel, text="Rol:").pack(anchor=tk.W)
        self.edit_role_entry = ttk.Entry(side_panel, width=25)
        self.edit_role_entry.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(side_panel, text="💾 Actualizar Info", command=self.update_member_info).pack(fill=tk.X, pady=(0, 5))
        ttk.Button(side_panel, text="🗑️ Eliminar Miembro", command=self.delete_member).pack(fill=tk.X)
        
        # Gestión de vacaciones
        ttk.Separator(side_panel, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=15)
        
        ttk.Label(side_panel, text="Gestionar vacaciones:", font=('Arial', 10, 'bold')).pack(anchor=tk.W, pady=(0, 10))
        
        ttk.Button(side_panel, text="📅 Seleccionar Fechas", command=self.select_vacation_dates).pack(fill=tk.X, pady=(0, 5))
        ttk.Button(side_panel, text="🔄 Actualizar Vista", command=self.refresh_grid).pack(fill=tk.X)
    
    def refresh_grid(self):
        """Actualiza el grid con los datos actuales"""
        # Limpiar grid existente
        for widget in self.grid_container.winfo_children():
            widget.destroy()
        
        # Headers
        header_style = {'background': '#4472C4', 'foreground': 'white', 'font': ('Arial', 9, 'bold')}
        
        tk.Label(self.grid_container, text="Nombre", width=15, **header_style).grid(row=0, column=0, sticky='nsew', padx=1, pady=1)
        tk.Label(self.grid_container, text="Rol", width=12, **header_style).grid(row=0, column=1, sticky='nsew', padx=1, pady=1)
        
        # Fechas como headers
        for col_idx, date in enumerate(self.planner.dates, start=2):
            date_text = date.strftime("%d\n%b")
            bg_color = '#E7E6E6' if self.planner.is_weekend(date) else '#4472C4'
            fg_color = 'black' if self.planner.is_weekend(date) else 'white'
            
            tk.Label(
                self.grid_container,
                text=date_text,
                width=6,
                background=bg_color,
                foreground=fg_color,
                font=('Arial', 8, 'bold')
            ).grid(row=0, column=col_idx, sticky='nsew', padx=1, pady=1)
        
        # Total column
        total_col = len(self.planner.dates) + 2
        tk.Label(self.grid_container, text="Total", width=8, **header_style).grid(row=0, column=total_col, sticky='nsew', padx=1, pady=1)
        
        # Datos de miembros
        for row_idx, member in enumerate(self.planner.team_members, start=1):
            # Nombre (clickable)
            name_label = tk.Label(
                self.grid_container,
                text=member.name,
                width=15,
                background='white',
                cursor='hand2',
                relief=tk.RAISED,
                borderwidth=1
            )
            name_label.grid(row=row_idx, column=0, sticky='nsew', padx=1, pady=1)
            name_label.bind('<Button-1>', lambda e, m=member: self.select_member(m))
            
            # Rol
            tk.Label(
                self.grid_container,
                text=member.role,
                width=12,
                background='white',
                relief=tk.RAISED,
                borderwidth=1
            ).grid(row=row_idx, column=1, sticky='nsew', padx=1, pady=1)
            
            # Días
            for col_idx, date in enumerate(self.planner.dates, start=2):
                date_str = date.strftime("%Y-%m-%d")
                is_vacation = member.is_on_vacation(date_str)
                is_weekend = self.planner.is_weekend(date)
                
                if is_vacation:
                    bg_color = '#92D050'
                    text = 'X'
                elif is_weekend:
                    bg_color = '#E7E6E6'
                    text = ''
                else:
                    bg_color = 'white'
                    text = ''
                
                cell = tk.Label(
                    self.grid_container,
                    text=text,
                    width=6,
                    background=bg_color,
                    relief=tk.RAISED,
                    borderwidth=1,
                    cursor='hand2'
                )
                cell.grid(row=row_idx, column=col_idx, sticky='nsew', padx=1, pady=1)
                cell.bind('<Button-1>', lambda e, m=member, d=date_str: self.toggle_vacation(m, d))
            
            # Total
            tk.Label(
                self.grid_container,
                text=str(len(member.vacation_days)),
                width=8,
                background='#F2F2F2',
                font=('Arial', 9, 'bold'),
                relief=tk.RAISED,
                borderwidth=1
            ).grid(row=row_idx, column=total_col, sticky='nsew', padx=1, pady=1)
    
    def select_member(self, member: TeamMember):
        """Selecciona un miembro para editar"""
        self.selected_member = member
        self.selected_label.config(text=member.name, foreground="blue")
        self.edit_name_entry.delete(0, tk.END)
        self.edit_name_entry.insert(0, member.name)
        self.edit_role_entry.delete(0, tk.END)
        self.edit_role_entry.insert(0, member.role)
    
    def toggle_vacation(self, member: TeamMember, date_str: str):
        """Alterna el estado de vacaciones de un día"""
        if member.is_on_vacation(date_str):
            member.remove_vacation_day(date_str)
        else:
            member.add_vacation_day(date_str)
        self.refresh_grid()
    
    def add_member(self):
        """Añade un nuevo miembro al equipo"""
        name = self.new_name_entry.get().strip()
        role = self.new_role_entry.get().strip()
        
        if not name:
            messagebox.showwarning("Advertencia", "El nombre es obligatorio")
            return
        
        if self.planner.get_member(name):
            messagebox.showwarning("Advertencia", f"Ya existe un miembro llamado {name}")
            return
        
        new_member = TeamMember(name, role)
        self.planner.add_member(new_member)
        
        self.new_name_entry.delete(0, tk.END)
        self.new_role_entry.delete(0, tk.END)
        
        self.refresh_grid()
        messagebox.showinfo("Éxito", f"Miembro {name} añadido correctamente")
    
    def update_member_info(self):
        """Actualiza la información de un miembro"""
        if not self.selected_member:
            messagebox.showwarning("Advertencia", "Selecciona un miembro primero")
            return
        
        new_name = self.edit_name_entry.get().strip()
        new_role = self.edit_role_entry.get().strip()
        
        if not new_name:
            messagebox.showwarning("Advertencia", "El nombre es obligatorio")
            return
        
        # Verificar si el nuevo nombre ya existe (y no es el mismo miembro)
        if new_name != self.selected_member.name and self.planner.get_member(new_name):
            messagebox.showwarning("Advertencia", f"Ya existe un miembro llamado {new_name}")
            return
        
        self.selected_member.name = new_name
        self.selected_member.role = new_role
        
        self.refresh_grid()
        self.select_member(self.selected_member)  # Mantener selección
        messagebox.showinfo("Éxito", "Información actualizada correctamente")
    
    def delete_member(self):
        """Elimina un miembro del equipo"""
        if not self.selected_member:
            messagebox.showwarning("Advertencia", "Selecciona un miembro primero")
            return
        
        if messagebox.askyesno("Confirmar", f"¿Eliminar a {self.selected_member.name}?"):
            self.planner.remove_member(self.selected_member.name)
            self.selected_member = None
            self.selected_label.config(text="Ninguno", foreground="gray")
            self.edit_name_entry.delete(0, tk.END)
            self.edit_role_entry.delete(0, tk.END)
            self.refresh_grid()
            messagebox.showinfo("Éxito", "Miembro eliminado correctamente")
    
    def select_vacation_dates(self):
        """Abre un diálogo para seleccionar fechas de vacaciones"""
        if not self.selected_member:
            messagebox.showwarning("Advertencia", "Selecciona un miembro primero")
            return
        
        # Crear ventana de diálogo
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Vacaciones de {self.selected_member.name}")
        dialog.geometry("400x500")
        
        ttk.Label(
            dialog,
            text=f"Selecciona las fechas de vacaciones para {self.selected_member.name}",
            font=('Arial', 10, 'bold')
        ).pack(pady=10)
        
        # Frame con scrollbar para las fechas
        frame = ttk.Frame(dialog)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        canvas = tk.Canvas(frame)
        scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor=tk.NW)
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Checkboxes para cada fecha
        date_vars = {}
        for date in self.planner.dates:
            date_str = date.strftime("%Y-%m-%d")
            var = tk.BooleanVar(value=self.selected_member.is_on_vacation(date_str))
            date_vars[date_str] = var
            
            date_label = date.strftime("%d de %B, %Y")
            if self.planner.is_weekend(date):
                date_label += " (Fin de semana)"
            
            cb = ttk.Checkbutton(scrollable_frame, text=date_label, variable=var)
            cb.pack(anchor=tk.W, pady=2)
        
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Botones
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)
        
        def apply_changes():
            # Actualizar días de vacaciones
            self.selected_member.vacation_days.clear()
            for date_str, var in date_vars.items():
                if var.get():
                    self.selected_member.add_vacation_day(date_str)
            self.refresh_grid()
            dialog.destroy()
            messagebox.showinfo("Éxito", "Vacaciones actualizadas correctamente")
        
        ttk.Button(btn_frame, text="Aplicar", command=apply_changes).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Cancelar", command=dialog.destroy).pack(side=tk.LEFT)
    
    def save_data(self):
        """Guarda los datos"""
        self.planner.save_data()
        messagebox.showinfo("Éxito", "Datos guardados correctamente")
    
    def export_excel(self):
        """Exporta a Excel"""
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile="vacaciones_navidad_2025.xlsx"
        )
        
        if filename:
            try:
                self.planner.export_to_excel(filename)
                messagebox.showinfo("Éxito", f"Exportado a Excel: {filename}")
            except Exception as e:
                messagebox.showerror("Error", f"Error al exportar: {str(e)}")
    
    def export_pdf(self):
        """Exporta a PDF"""
        filename = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
            initialfile="vacaciones_navidad_2025.pdf"
        )
        
        if filename:
            try:
                self.planner.export_to_pdf(filename)
                messagebox.showinfo("Éxito", f"Exportado a PDF: {filename}")
            except Exception as e:
                messagebox.showerror("Error", f"Error al exportar: {str(e)}")


def main():
    """Función principal"""
    root = tk.Tk()
    app = HolidayPlannerGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
